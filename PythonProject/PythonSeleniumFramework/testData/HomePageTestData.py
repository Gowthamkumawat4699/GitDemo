
import openpyxl

class HomePageData:

    home_page_data = [{"First_Name": "Rahul","Last_Name": "Shetty","Email": "Rahalshetty@gmail.com","Gender": "Male","Password":"Rahul@1234"},{"First_Name":"Anshika","Last_Name":"Shetty","Email":"AnshikaShetty@gmail.com","Gender":"Female","Password":"Anshika123"}]

    @staticmethod
    def get_test_data(test_case_name):
        file = openpyxl.load_workbook("C:\\Users\\91709\\OneDrive\\Documents\\TestData.xlsx")
        sheet = file.active
        data = {}
        for i in range(1 , sheet.max_row + 1) :
            if sheet.cell(row=i , column=1).value == test_case_name :
                for j in range(2 , sheet.max_column + 1) :
                    data[sheet.cell(row=1 , column=j).value] = sheet.cell(row=i , column=j).value
        print("DATA: ",data)
        print("Lists DATA: " , [data])
        return [data]
