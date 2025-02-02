import openpyxl

file = openpyxl.load_workbook("C:\\Users\\91709\\OneDrive\\Documents\\TestData.xlsx")
sheet = file.active

print(sheet.cell(row=2,column=2).value)


data = {}
for i in range(1,sheet.max_row+1):
     if sheet.cell(row=i,column=1).value == 'Test One':
        for j in range(2,sheet.max_column+1):
            data[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value


print(data)