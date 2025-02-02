import os
import time

import openpyxl
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
filepath = "C:\\Users\\91709\\Downloads\\download.xlsx"
driver = webdriver.Chrome()
driver.get('http://rahulshettyacademy.com/upload-download-test/index.html')
driver.maximize_window()
time.sleep(5)

if os.path.exists(filepath):
    os.remove(filepath)

driver.find_element(By.ID,"downloadButton").click()

driver.find_element(By.ID, "downloadButton")
driver.implicitly_wait(5)
time.sleep(3)
workbook = openpyxl.load_workbook(filepath)
sheet = workbook.active
dict = {}
print(sheet.cell(row=1,column=2).value)

for i in range(1,sheet.max_column+1):
       if sheet.cell(row=1,column=i).value == 'price':
           dict["col"] = i
           break


for r in range(1,sheet.max_row+1):
    for c in range(1,sheet.max_column+1):
       if sheet.cell(row=r,column=c).value == 'Apple':
           dict["row"] = r
           break


sheet.cell(row=dict['row'],column=dict['col']).value = 100
time.sleep(3)
workbook.save(filepath)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(filepath)
wait = WebDriverWait(driver,5)
wait.until(ec.visibility_of_element_located((By.XPATH, "//div[@role='alert']/div[2]")))
print(driver.find_element(By.XPATH, "//div[@role='alert']/div[2]").text)
#Smart xpath concept
fruitName = "Apple"
price_col_id = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
price = driver.find_element(By.XPATH, "//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+price_col_id+"-undefined']").text
print(f"{fruitName}: {price}")



